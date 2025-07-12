import openpyxl 
from datetime import datetime
import os

class Product:
    def __init__(self, product_id, name, price, stock):
        self.product_id = product_id
        self.name = name
        self.price = float(price)
        self.stock = int(stock)

    def update_stock(self, quantity):
        if self.stock >= quantity:
            self.stock -= quantity
            return True
        return False

    def to_dict(self):
        return {
            'product_id': self.product_id,
            'name': self.name,
            'price': self.price,
            'stock': self.stock
        }

class Customer:
    def __init__(self, customer_id, name, email):
        self.customer_id = customer_id
        self.name = name
        self.email = email
        self.discount = 0

    def to_dict(self):
        return {
            'customer_id': self.customer_id,
            'name': self.name,
            'email': self.email
        }

class PremiumCustomer(Customer):
    def __init__(self, customer_id, name, email):
        super().__init__(customer_id, name, email)
        self.discount = 0.1  # 10% discount

class Order:
    def __init__(self, order_id, customer, products, quantities):
        self.order_id = order_id
        self.customer = customer
        self.products = products
        self.quantities = quantities
        self.total_amount = self.calculate_total()
        self.date = datetime.now()

    def calculate_total(self):
        total = sum(p.price * q for p, q in zip(self.products, self.quantities))
        return total * (1 - self.customer.discount)

    def to_dict(self):
        return {
            'order_id': self.order_id,
            'customer_name': self.customer.name,
            'products': [p.name for p in self.products],
            'total_amount': self.total_amount,
            'date': self.date
        }

class ECommerceSystem:
    def __init__(self):
        self.products = {}
        self.customers = {}
        self.orders = {}
        self.load_data()

    def load_data(self):
        if os.path.exists('products.xlsx'):
            wb = openpyxl.load_workbook('products.xlsx')
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Check if product_id exists
                    self.products[row[0]] = Product(row[0], row[1], row[2], row[3])

        if os.path.exists('customers.xlsx'):
            wb = openpyxl.load_workbook('customers.xlsx')
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Check if customer_id exists
                    if row[3] == 'Premium':
                        self.customers[row[0]] = PremiumCustomer(row[0], row[1], row[2])
                    else:
                        self.customers[row[0]] = Customer(row[0], row[1], row[2])

    def save_data(self):
        try:
            # Save products
            print("Saving products...")
            print(f"Current products in memory: {len(self.products)}")
            print(f"Current working directory: {os.getcwd()}")
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['Product ID', 'Name', 'Price', 'Stock'])
            for product in self.products.values():
                print(f"Saving product: {product.product_id}")
                ws.append([product.product_id, product.name, product.price, product.stock])
            
            file_path = os.path.join(os.getcwd(), 'products.xlsx')
            print(f"Attempting to save to: {file_path}")
            wb.save(file_path)
            print("Products saved successfully!")

            # Save customers
            print("Saving customers...")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['Customer ID', 'Name', 'Email', 'Type'])
            for customer in self.customers.values():
                customer_type = 'Premium' if isinstance(customer, PremiumCustomer) else 'Regular'
                ws.append([customer.customer_id, customer.name, customer.email, customer_type])
            wb.save('customers.xlsx')
            print("Customers saved successfully!")
        except Exception as e:
            print(f"Error saving data: {str(e)}")
            import traceback
            traceback.print_exc()

def main():
    system = ECommerceSystem()
    
    while True:
        print("\nWelcome to the E-Commerce System!")
        print("1. Add a new product")
        print("2. Register a new customer")
        print("3. Place an order")
        print("4. View order details")
        print("5. Exit")
        
        choice = input("Choose an option: ")

        if choice == '1':
            product_id = input("Enter Product ID: ")
            name = input("Enter Product Name: ")
            price = float(input("Enter Product Price: "))
            stock = int(input("Enter Product Stock: "))
            system.products[product_id] = Product(product_id, name, price, stock)
            print("Product added successfully!")

        elif choice == '2':
            customer_id = input("Enter Customer ID: ")
            name = input("Enter Customer Name: ")
            email = input("Enter Customer Email: ")
            is_premium = input("Is this a premium customer? (y/n): ").lower() == 'y'
            if is_premium:
                system.customers[customer_id] = PremiumCustomer(customer_id, name, email)
            else:
                system.customers[customer_id] = Customer(customer_id, name, email)
            print("Customer registered successfully!")

        elif choice == '3':
            customer_id = input("Enter Customer ID: ")
            if customer_id not in system.customers:
                print("Customer not found!")
                continue

            products = []
            quantities = []
            while True:
                product_id = input("Enter Product ID (or 'done' to finish): ")
                if product_id.lower() == 'done':
                    break
                if product_id not in system.products:
                    print("Product not found!")
                    continue
                quantity = int(input("Enter Quantity: "))
                if not system.products[product_id].update_stock(quantity):
                    print("Insufficient stock!")
                    continue
                products.append(system.products[product_id])
                quantities.append(quantity)

            if products:
                order_id = f"O{len(system.orders) + 1:03d}"
                order = Order(order_id, system.customers[customer_id], products, quantities)
                system.orders[order_id] = order
                print("Order placed successfully!")

        elif choice == '4':
            if not system.orders:
                print("No orders to display!")
                continue
            for order in system.orders.values():
                print("\nOrder Details:")
                print(f"Order ID: {order.order_id}")
                print(f"Customer: {order.customer.name}")
                print(f"Products: {[p.name for p in order.products]}")
                print(f"Total Amount: ${order.total_amount:.2f}")

        elif choice == '5':
            print("Saving data before exit...")
            system.save_data()
            print("Save completed. Exiting... Goodbye!")
            break

if __name__ == "__main__":
    main()
